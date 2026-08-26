import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Генератор отчетов Статусы", layout="centered")
st.title("📊 Генератор отчета «Статусы»")
st.write("Загрузите файл «Выгрузка.xlsx» и получите отчет «Статусы»")

uploaded_file = st.file_uploader("Выберите файл выгрузки (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    # Читаем данные
    df = pd.read_excel(uploaded_file, header=0)
    
    with st.expander("👀 Превью загруженных данных"):
        st.dataframe(df.head(10))
    
    # Находим нужные колонки
    status_col = None
    responsible_col = None
    date_col = None
    
    for col in df.columns:
        col_str = str(col).strip()
        if "Стадия" in col_str:
            status_col = col
        if "Ответственный" in col_str:
            responsible_col = col
        if "Дата изменения" in col_str:
            date_col = col
    
    if status_col is None or responsible_col is None:
        st.error(f"❌ Не найдены нужные колонки. Найдены: {list(df.columns)}")
        st.stop()
    
    # Создаем копию с нужными колонками
    df_clean = df[[status_col, responsible_col]].copy()
    if date_col is not None:
        df_clean["Дата изменения"] = df[date_col]
    
    df_clean.columns = ["Стадия", "Ответственный"] + (["Дата изменения"] if date_col is not None else [])
    
    # Объединяем статусы
    df_clean["Стадия"] = df_clean["Стадия"].replace(
        {
            "Аккаунты_Менеджер назначен": "Менеджер назначен",
            "Аккаунты_Менеджер назначен ": "Менеджер назначен",
        }
    )
    
    # Строим сводную таблицу
    pivot = pd.crosstab(df_clean["Стадия"], df_clean["Ответственный"])
    
    # Добавляем строку "Без изменений >10 дней"
    if date_col is not None:
        df_clean["Дата изменения"] = pd.to_datetime(df_clean["Дата изменения"], errors="coerce")
        today = datetime.now()
        
        old_leads = df_clean[
            (df_clean["Дата изменения"].notna()) & 
            ((today - df_clean["Дата изменения"]) > timedelta(days=10))
        ]
        
        old_count = old_leads.groupby("Ответственный").size()
        pivot.loc["Без изменений >10 дней"] = old_count.reindex(pivot.columns, fill_value=0)
    
    # Добавляем столбец "Итого"
    pivot["Итого"] = pivot.sum(axis=1)
    
    # Добавляем строку "Итого"
    total_row = pivot.sum(axis=0)
    pivot.loc["Итого"] = total_row
    
    pivot.index.name = "Статус лида"
    
    st.subheader("📋 Сгенерированный отчет")
    st.dataframe(pivot, use_container_width=True)
    
    # --- СОЗДАНИЕ EXCEL С ФОРМАТИРОВАНИЕМ ---
    output = BytesIO()
    
    temp_output = BytesIO()
    with pd.ExcelWriter(temp_output, engine="openpyxl") as writer:
        pivot.to_excel(writer, sheet_name="Статусы")
    
    from openpyxl import load_workbook
    
    temp_output.seek(0)
    wb = load_workbook(temp_output)
    ws = wb["Статусы"]
    
    # --- ШИРИНА КОЛОНОК ---
    # Колонка A — 37
    ws.column_dimensions["A"].width = 37
    # Остальные колонки (B, C, D, E, F) — 18
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 18
    
    # --- ЦВЕТА ---
    # Цвет шапки (светло-песочный)
    color_header = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
    # Цвет для строк: Менеджер назначен, Пора звонить, Пора звонить (холодняк)
    color_highlight = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    # Цвет для строки Итого
    color_total = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    
    # Жирный шрифт для шапки и строки Итого
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # --- 1. ЗАКРЕПЛЯЕМ ШАПКУ (первая строка) ---
    ws.freeze_panes = "A2"  # Закрепляем первую строку
    
    # --- 2. ФОРМАТИРУЕМ ШАПКУ (первая строка) ---
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = color_header
        cell.font = bold_font
        cell.alignment = center_alignment
    
    # --- 3. ФОРМАТИРУЕМ ОСТАЛЬНЫЕ СТРОКИ ---
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        
        # Строка "Итого"
        if cell_value == "Итого":
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = color_total
                cell.font = bold_font
                cell.alignment = center_alignment
        
        # Строки: Менеджер назначен, Пора звонить, Пора звонить (холодняк)
        elif cell_value in ["Менеджер назначен", "Пора звонить", "Пора звонить (холодняк)"]:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = color_highlight
                cell.alignment = center_alignment
        
        # Остальные строки — выравниваем по центру только колонки B-F
        else:
            for col in range(2, ws.max_column + 1):
                ws.cell(row=row, column=col).alignment = center_alignment
    
    wb.save(output)
    output.seek(0)
    
    st.download_button(
        label="⬇️ Скачать отчет Статусы.xlsx",
        data=output,
        file_name="Статусы.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.success("✅ Готово! Нажмите кнопку выше, чтобы скачать файл.")
